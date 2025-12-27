import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import os
import re

# =====================
# CONFIG
# =====================
# Empty by default - can be customized by user
IGNORE_COLUMNS = set()

# =====================
# STYLES
# =====================
MODIFIED_FILL = PatternFill("solid", fgColor="FFF59D")
ADDED_FILL = PatternFill("solid", fgColor="C8E6C9")
REMOVED_FILL = PatternFill("solid", fgColor="FFCDD2")

# =====================
# UI APPLICATION
# =====================
class ExcelComparerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Comparison Tool")
        self.root.geometry("700x400")

        self.file1 = None
        self.file2 = None
        self.output_dir = None

        self.build_ui()

    def build_ui(self):
        tk.Button(self.root, text="Upload Excel File 1", command=self.load_file1).pack(pady=5)
        self.file1_label = tk.Label(self.root, text="")
        self.file1_label.pack()

        tk.Button(self.root, text="Upload Excel File 2", command=self.load_file2).pack(pady=5)
        self.file2_label = tk.Label(self.root, text="")
        self.file2_label.pack()

        tk.Button(self.root, text="Select Output Folder", command=self.load_output).pack(pady=5)
        self.output_label = tk.Label(self.root, text="")
        self.output_label.pack()

        self.compare_btn = tk.Button(self.root, text="Compare Excel Files", command=self.compare)
        self.compare_btn.pack(pady=15)

        self.clear_btn = tk.Button(self.root, text="Clear Selection", command=self.clear_selection)
        self.clear_btn.pack(pady=5)

        self.status_label = tk.Label(self.root, text="", fg="blue")
        self.status_label.pack()

    def load_file1(self):
        self.file1 = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
        self.file1_label.config(text=self.file1 or "")

    def load_file2(self):
        self.file2 = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
        self.file2_label.config(text=self.file2 or "")

    def load_output(self):
        self.output_dir = filedialog.askdirectory()
        self.output_label.config(text=self.output_dir or "")

    def clear_selection(self):
        self.file1 = None
        self.file2 = None
        self.output_dir = None

        self.file1_label.config(text="")
        self.file2_label.config(text="")
        self.output_label.config(text="")
        self.status_label.config(text="", fg="blue")

    
    def normalize_val(self, val):
        # Handle None/NaN values - treat as empty string
        if pd.isna(val) or val is None:
            return ""
        
        if isinstance(val, str):
            # Remove all Unicode whitespace characters (including non-breaking spaces, etc.)
            # First, remove zero-width characters and other invisible Unicode
            # Remove zero-width space, zero-width non-joiner, zero-width joiner, etc.
            val = re.sub(r'[\u200B-\u200D\uFEFF]', '', val)  # Remove zero-width characters
            
            # Remove all types of whitespace (spaces, tabs, newlines, non-breaking spaces, etc.)
            # First strip leading/trailing whitespace (including \n, \r, \r\n, non-breaking spaces at start/end)
            stripped = val.strip()  # Removes leading/trailing whitespace including newlines
            
            # Replace all Unicode whitespace (including non-breaking space \u00A0) with regular space
            # Then normalize multiple spaces to single space
            normalized = re.sub(r'[\s\u00A0]+', ' ', stripped)  # Normalize all whitespace to single space
            
            # If after normalization it's empty or only whitespace, return empty string
            if not normalized or normalized.isspace():
                return ""
            
            return normalized.lower()
        
        # For numeric types, convert to string for consistent comparison
        if isinstance(val, (int, float)):
            # For floats, remove trailing zeros
            if isinstance(val, float):
                str_val = f"{val:.10f}".rstrip('0').rstrip('.')
                return str_val if str_val else "0"
            return str(val)
        
        # For other types, convert to string and normalize
        str_val = str(val).strip()
        if not str_val or str_val.isspace():
            return ""
        return re.sub(r'\s+', ' ', str_val).lower()

    def compare(self):
        if not all([self.file1, self.file2, self.output_dir]):
            self.status_label.config(text="Please upload both files and select output folder.", fg="red")
            return

        self.status_label.config(text="Comparing files... Please wait ⏳", fg="blue")
        self.root.update_idletasks()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(
            self.output_dir,
            f"{timestamp}_excel_differences.xlsx"
        )

        wb = Workbook()
        wb.remove(wb.active)

        summary_data = []

        xls1 = pd.read_excel(self.file1, sheet_name=None)
        xls2 = pd.read_excel(self.file2, sheet_name=None)

        common_sheets = set(xls1.keys()).intersection(xls2.keys())

        for sheet in common_sheets:
            df1 = xls1[sheet].drop(columns=IGNORE_COLUMNS, errors="ignore").fillna("")
            df2 = xls2[sheet].drop(columns=IGNORE_COLUMNS, errors="ignore").fillna("")

            # Dynamically detect if key-based matching would be better
            # Look for identifier-like columns that could serve as keys
            def detect_key_columns(df1, df2):
                """Detect potential key columns for matching"""
                # Generic identifier patterns
                possible_key_patterns = [
                    'id', 'code', 'key', 'identifier', 'pk', 'primary_key',
                    'name', 'title', 'label', 'sku', 'product_id', 'item_id'
                ]
                key_columns = []
                
                for col in df1.columns:
                    if col not in df2.columns:
                        continue
                    col_lower = col.lower()
                    # Check if column name suggests it's an identifier
                    if any(pattern in col_lower for pattern in possible_key_patterns):
                        # Check if this column has mostly unique values (good for key)
                        unique_ratio_df1 = df1[col].nunique() / len(df1) if len(df1) > 0 else 0
                        unique_ratio_df2 = df2[col].nunique() / len(df2) if len(df2) > 0 else 0
                        # If at least 50% unique, it's a good candidate
                        if unique_ratio_df1 > 0.5 and unique_ratio_df2 > 0.5:
                            key_columns.append(col)
                
                return key_columns
            
            key_columns = detect_key_columns(df1, df2)
            use_key_matching = len(key_columns) > 0
            
            # Helper function to create normalized row signature for comparison
            def get_row_signature(row, use_keys=False):
                if use_keys and key_columns:
                    # Use key columns for matching
                    return tuple(self.normalize_val(row[col]) for col in key_columns)
                else:
                    # Full row matching
                    return tuple(self.normalize_val(val) for val in row.values)

            # Try LCS first, then check if key-based matching would be better
            added = removed = modified = 0
            differences = []
            
            # First, try standard LCS approach
            def run_lcs_comparison():
                """Run LCS comparison and return differences"""
                rows1_sig = [tuple(self.normalize_val(val) for val in df1.iloc[i].values) for i in range(len(df1))]
                rows2_sig = [tuple(self.normalize_val(val) for val in df2.iloc[i].values) for i in range(len(df2))]
                
                m, n = len(rows1_sig), len(rows2_sig)
                dp = [[0] * (n + 1) for _ in range(m + 1)]
                
                for i in range(1, m + 1):
                    for j in range(1, n + 1):
                        if rows1_sig[i-1] == rows2_sig[j-1]:
                            dp[i][j] = dp[i-1][j-1] + 1
                        else:
                            dp[i][j] = max(dp[i-1][j], dp[i][j-1])
                
                matched_df1 = set()
                matched_df2 = set()
                diffs = []
                a = r = 0
                
                i, j = m, n
                while i > 0 or j > 0:
                    if i > 0 and j > 0 and rows1_sig[i-1] == rows2_sig[j-1]:
                        matched_df1.add(i-1)
                        matched_df2.add(j-1)
                        i -= 1
                        j -= 1
                    elif j > 0 and (i == 0 or dp[i][j-1] >= dp[i-1][j]):
                        diffs.insert(0, ("added", df2.iloc[j-1]))
                        a += 1
                        j -= 1
                    elif i > 0 and (j == 0 or dp[i][j-1] < dp[i-1][j]):
                        diffs.insert(0, ("removed", df1.iloc[i-1]))
                        r += 1
                        i -= 1
                
                return diffs, a, r, matched_df1, matched_df2
            
            # Run LCS first
            lcs_differences, lcs_added, lcs_removed, lcs_matched_df1, lcs_matched_df2 = run_lcs_comparison()
            
            # Check if key-based matching might be better
            # If LCS shows many differences but we have key columns, try key-based matching
            total_lcs_diffs = lcs_added + lcs_removed
            use_key_fallback = False
            
            if use_key_matching and key_columns and total_lcs_diffs > min(len(df1), len(df2)) * 0.3:
                # LCS shows >30% differences - try key-based matching
                # Check if key-based matching produces fewer differences
                use_key_fallback = True
            
            if use_key_matching and key_columns and use_key_fallback:
                # Match by key, then identify modifications
                # Build maps of key signature to row indices
                key_sig_to_df1 = {}
                key_sig_to_df2 = {}
                
                for idx in range(len(df1)):
                    key_sig = get_row_signature(df1.iloc[idx], use_keys=True)
                    if key_sig not in key_sig_to_df1:
                        key_sig_to_df1[key_sig] = []
                    key_sig_to_df1[key_sig].append(idx)
                
                for idx in range(len(df2)):
                    key_sig = get_row_signature(df2.iloc[idx], use_keys=True)
                    if key_sig not in key_sig_to_df2:
                        key_sig_to_df2[key_sig] = []
                    key_sig_to_df2[key_sig].append(idx)
                
                # Track matched rows
                matched_df1_indices = set()
                matched_df2_indices = set()
                
                # Process df2 in order to maintain output order
                for idx2 in range(len(df2)):
                    if idx2 in matched_df2_indices:
                        continue
                    
                    row2 = df2.iloc[idx2]
                    key_sig2 = get_row_signature(row2, use_keys=True)
                    
                    # Check if this key exists in df1
                    if key_sig2 in key_sig_to_df1 and key_sig_to_df1[key_sig2]:
                        # Match found - check if it's modified
                        idx1 = key_sig_to_df1[key_sig2].pop(0)
                        matched_df1_indices.add(idx1)
                        matched_df2_indices.add(idx2)
                        
                        row1 = df1.iloc[idx1]
                        
                        # Compare full rows to check for modifications
                        full_sig1 = tuple(self.normalize_val(val) for val in row1.values)
                        full_sig2 = tuple(self.normalize_val(val) for val in row2.values)
                        
                        if full_sig1 != full_sig2:
                            # Modified row
                            differences.append(("modified", row1, row2))
                            modified += 1
                        # If they match fully, skip (unchanged)
                    else:
                        # Key doesn't exist in df1 - added row
                        differences.append(("added", row2))
                        added += 1
                        matched_df2_indices.add(idx2)
                
                # Any unmatched rows in df1 are removed
                for idx1 in range(len(df1)):
                    if idx1 not in matched_df1_indices:
                        differences.append(("removed", df1.iloc[idx1]))
                        removed += 1
            else:
                # Use LCS results (already computed)
                differences = lcs_differences
                added = lcs_added
                removed = lcs_removed
            

            # Count actual differences that will be written
            actual_added = sum(1 for d in differences if d[0] == "added")
            actual_removed = sum(1 for d in differences if d[0] == "removed")
            actual_modified = sum(1 for d in differences if d[0] == "modified")
            
            # Only create sheet if there are differences
            if actual_added > 0 or actual_removed > 0 or actual_modified > 0:
                ws = wb.create_sheet(sheet)
                ws.append(list(df1.columns))

                # Write differences to sheet
                for diff in differences:
                    diff_type = diff[0]
                    if diff_type == "added":
                        row = diff[1]
                        ws.append(list(row))
                        for cell in ws[ws.max_row]:
                            cell.fill = ADDED_FILL
                    elif diff_type == "removed":
                        row = diff[1]
                        ws.append(list(row))
                        for cell in ws[ws.max_row]:
                            cell.fill = REMOVED_FILL
                    elif diff_type == "modified":
                        row1 = diff[1]
                        row2 = diff[2]
                        ws.append(list(row2))
                        excel_row = ws.max_row
                        for col_idx, col in enumerate(df1.columns, start=1):
                            val1 = self.normalize_val(row1[col])
                            val2 = self.normalize_val(row2[col])
                            if val1 != val2:
                                ws.cell(row=excel_row, column=col_idx).fill = MODIFIED_FILL

                summary_data.append([
                    sheet,
                    len(df1),
                    len(df2),
                    actual_added,
                    actual_removed,
                    actual_modified,
                    actual_added + actual_removed + actual_modified
                ])

        # =====================
        # SUMMARY SHEET
        # =====================
        summary_ws = wb.create_sheet("Difference_Summary", 0)
        summary_ws.append([
            "Sheet Name",
            "File 1 Rows",
            "File 2 Rows",
            "Rows Added",
            "Rows Removed",
            "Rows Modified",
            "Total Differences",
            "Validation"
        ])

        for row in summary_data:
            # Validate: File1 + Added - Removed = File2
            file1_count = row[1]
            file2_count = row[2]
            added_count = row[3]
            removed_count = row[4]
            
            expected_file2 = file1_count + added_count - removed_count
            validation = "✓ OK" if expected_file2 == file2_count else f"✗ MISMATCH (expected {expected_file2})"
            
            summary_ws.append(row + [validation])

        wb.save(output_file)

        self.status_label.config(
            text=f"Comparison completed ✔\nOutput saved at:\n{output_file}",
            fg="green"
        )


# =====================
# RUN APP
# =====================
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparerApp(root)
    root.mainloop()
